"""Export everything in mhp.db into one organized, easy-to-use package.

Produces ~/Projects/mhp-brain/export/:
  MHP_Data_Master.xlsx   one workbook, Overview tab + a tab per table (open this first)
  csv/<table>.csv        plain CSV per table (portable, opens anywhere)
  reports/*.md           the generated analysis reports, copied in so it's self-contained
  README.md              index explaining what's what

Read-only on mhp.db (mode=ro). Re-runnable — overwrites the export folder's generated files.

Run:  python3 export.py
"""
import csv
import shutil
import sqlite3
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
DB = HERE / "mhp.db"
OUT = HERE / "export"

# Tables to export, in human order, with a friendly tab/section name. Empty tables are noted in
# the Overview but get no tab.
TABLES = [
    ("projects",   "Projects"),
    ("estimates",  "Estimates"),
    ("line_items", "Line Items"),
    ("unit_costs", "Unit Cost Catalog"),
    ("lump_costs", "Lump-Sum Catalog"),
    ("actuals",    "Actuals"),
    ("subs",       "Subcontractors"),
    ("crew",       "Crew"),
]

# Columns that hold dollars, per table — formatted as currency in the workbook.
MONEY = {
    "estimates": {"sum_item_total", "sum_sov_total", "stated_total"},
    "line_items": {"unit_price", "material", "labor", "sub_bid", "item_total", "sov_total"},
    "unit_costs": {"median_unit_price", "p25", "p75", "min_price", "max_price"},
    "lump_costs": {"median_total", "p25", "p75"},
    "actuals": {"closing_total"},
    "crew": {"rate"},
}

REPORTS = ["cost_catalog.md", "ANALYSIS.md", "PROFIT_OPPORTUNITIES.md",
           "parse_report.md", "variance.md"]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # dark slate, no gradient
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16)
SUB_FONT = Font(italic=True, color="6B7280")


def pretty(col):
    """projects.last_activity -> 'Last Activity', with a few overrides."""
    fix = {"id": "ID", "p25": "P25", "p75": "P75", "sov_total": "SOV Total",
           "n_jobs": "# Jobs", "n_lines": "# Lines", "item_no": "Item #",
           "qty": "Qty", "est_date": "Est. Date", "canon_desc": "Canonical Desc"}
    if col in fix:
        return fix[col]
    return col.replace("_", " ").title()


def cols(con, table):
    return [r[1] for r in con.execute(f"PRAGMA table_info({table})")]


def write_sheet(ws, con, table):
    columns = cols(con, table)
    rows = con.execute(f"SELECT {', '.join(columns)} FROM {table}").fetchall()
    money = MONEY.get(table, set())

    # header
    for c, col in enumerate(columns, 1):
        cell = ws.cell(1, c, pretty(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    # data
    for r, row in enumerate(rows, 2):
        for c, (col, val) in enumerate(zip(columns, row), 1):
            cell = ws.cell(r, c, val)
            if col in money and isinstance(val, (int, float)):
                cell.number_format = '$#,##0.00'
    # widths (capped) + freeze header
    for c, col in enumerate(columns, 1):
        sample = [str(pretty(col))] + [str(rw[c - 1]) for rw in rows[:200] if rw[c - 1] is not None]
        width = min(max((len(s) for s in sample), default=10) + 2, 48)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    return len(rows)


def write_overview(ws, con, counts, today):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws["A1"] = "MHP Brain — Data Export"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"All estimate data captured so far. Generated {today}."
    ws["A2"].font = SUB_FONT

    ws["A4"] = "What's in this workbook"
    ws["A4"].font = Font(bold=True, size=12)
    guide = [
        ("Projects", "Every job MHP has touched — type, market, status, last activity."),
        ("Estimates", "Each estimate file parsed: line count, bid total, parse confidence, flags."),
        ("Line Items", "Every priced line across all estimates — the raw detail."),
        ("Unit Cost Catalog", "Proven per-unit rates (median + p25/p75 band) from real bids."),
        ("Lump-Sum Catalog", "Proven lump-sum item totals (cabinets, general conditions, etc.)."),
        ("Actuals", "Closeout totals captured so far (sparse — fills in with QuickBooks)."),
        ("Subcontractors", "Sub roster — trade, phone, jobs worked."),
        ("Crew", "Company crew directory — role, rate, contact."),
    ]
    r = 5
    for name, desc in guide:
        ws.cell(r, 1, name).font = Font(bold=True)
        ws.cell(r, 2, desc)
        r += 1

    r += 1
    ws.cell(r, 1, "Counts").font = Font(bold=True, size=12)
    r += 1
    for table, label in TABLES:
        ws.cell(r, 1, label)
        ws.cell(r, 2, f"{counts[table]:,} rows")
        r += 1

    r += 1
    ws.cell(r, 1, "Also in this folder").font = Font(bold=True, size=12)
    r += 1
    for txt in ["csv/ — every table as a plain .csv (opens in anything)",
                "reports/ — the analysis write-ups (pricing, catalog, parse, variance)",
                "README.md — this guide in text form"]:
        ws.cell(r, 2, txt)
        r += 1


def main():
    con = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    today = date.today().isoformat()

    OUT.mkdir(exist_ok=True)
    (OUT / "csv").mkdir(exist_ok=True)
    (OUT / "reports").mkdir(exist_ok=True)

    counts = {t: con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0] for t, _ in TABLES}

    # --- master workbook ---
    wb = Workbook()
    write_overview(wb.active, con, counts, today)
    wb.active.title = "Overview"
    for table, label in TABLES:
        ws = wb.create_sheet(label[:31])
        write_sheet(ws, con, table)
    xlsx = OUT / "MHP_Data_Master.xlsx"
    wb.save(xlsx)

    # --- per-table CSVs ---
    for table, _ in TABLES:
        columns = cols(con, table)
        with open(OUT / "csv" / f"{table}.csv", "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(columns)
            w.writerows(con.execute(f"SELECT {', '.join(columns)} FROM {table}"))

    # --- bundle reports ---
    copied = []
    for name in REPORTS:
        src = HERE / name
        if src.exists():
            shutil.copy2(src, OUT / "reports" / name)
            copied.append(name)

    # --- README index ---
    lines = [f"# MHP Brain — Data Export ({today})", "",
             "All MHP estimate data captured so far, organized for easy use.", "",
             "## Start here", "",
             "**`MHP_Data_Master.xlsx`** — one workbook, a tab per table. Open this first.", "",
             "## Tables", "",
             "| Table | Rows | What it is |", "|---|--:|---|"]
    desc = {"projects": "Every job — type, market, status, last activity",
            "estimates": "Each parsed estimate — totals, confidence, flags",
            "line_items": "Every priced line across all estimates (raw detail)",
            "unit_costs": "Proven per-unit rates (median + p25/p75)",
            "lump_costs": "Proven lump-sum item totals",
            "actuals": "Closeout totals so far (sparse)",
            "subs": "Subcontractor roster",
            "crew": "Company crew directory"}
    for table, label in TABLES:
        lines.append(f"| {label} | {counts[table]:,} | {desc[table]} |")
    lines += ["", "## Folders", "",
              "- **`csv/`** — every table as plain `.csv`",
              "- **`reports/`** — analysis write-ups: " + ", ".join(f"`{c}`" for c in copied),
              "", "---", "",
              "*Read-only export from `mhp.db`. Re-run `python3 export.py` to refresh.*"]
    (OUT / "README.md").write_text("\n".join(lines) + "\n")

    con.close()
    total = sum(counts.values())
    print(f"Exported {total:,} rows across {len(TABLES)} tables -> {OUT}/")
    print(f"  {xlsx.name} ({xlsx.stat().st_size//1024} KB), "
          f"{len(TABLES)} CSVs, {len(copied)} reports, README.md")


if __name__ == "__main__":
    main()
