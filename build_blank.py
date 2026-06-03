"""Create a reusable blank MHP estimate template (formulas intact, inputs cleared).

The engine writes generated estimates into a copy of this, so output is a native MHP .xlsx
the estimator opens in Excel and the Item Total / Markup / SOV formulas recompute on open.
"""
import shutil
from pathlib import Path

import openpyxl

SRC = Path.home() / ("Desktop/Walt/Clients/MHP Construction (Josh)/01_Active Projects/"
                     "Michael Mason - Kitchen Reno/Mason Kitchen Reno Project/"
                     "Michael Mason Kitchen Project - Current/Mason Kitchen Estimate 11-10-25.xlsx")
OUT = Path(__file__).parent / "mhp_template_blank.xlsx"

ITEM_ROWS = range(7, 183)        # CSI item rows
INPUT_COLS = [3, 5, 6, 7, 8]     # C=qty E=unit_price F=material G=labor H=sub_bid


def main():
    wb = openpyxl.load_workbook(str(SRC), data_only=False)   # keep formulas
    ws = wb["Template"]
    cleared = 0
    for r in ITEM_ROWS:
        if ws.cell(row=r, column=1).value is None and ws.cell(row=r, column=2).value is None:
            continue
        for c in INPUT_COLS:
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None
                cleared += 1
    # drop the canned example + scratch sheets so the file is clean
    for junk in ("Filled In Example", "Sheet3"):
        if junk in wb.sheetnames:
            del wb[junk]
    wb.save(str(OUT))
    print(f"Wrote {OUT.name} — cleared {cleared} input cells, formulas preserved")


if __name__ == "__main__":
    main()
