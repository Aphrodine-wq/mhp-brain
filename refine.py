"""Data-quality pass: real subcontractor roster + honest project status.

Subs: the mined CSV only caught 10 (filename mentions). Real subs live in (a) the Corinth
subcontractor list, (b) the 'Subcontractor Name' column inside estimate sheets, (c) the old CSV.
Projects: 'Active' was just folder location. Re-bucket by latest document activity (today 2026-06)
and pull in the _BID PIPELINE / _PAUSED jobs that weren't in the index at all.
"""
import csv
import re
import sqlite3
from pathlib import Path

import openpyxl

MHP = Path.home() / "Desktop/Walt/Clients/MHP Construction (Josh)"
DB = Path(__file__).parent / "mhp.db"
TODAY_YEAR = "2026"

NON_NAMES = {"", "name", "subcontractor name", "n/a", "na", "tbd", "self", "mhp", "various",
             "self performed", "in house", "0"}
NOTE_WORDS = ("allowance", "sqft", "sq ft", "option", "included", "is $", "per ", "approx",
              "estimate", " each", "material", "labor", "/ft", "/sf", "total")


def looks_like_name(s):
    """Reject notes/allowances typed into the sub column ('150 for fiberglass door')."""
    s = cname(s)
    if not s or len(s) > 38 or "$" in s or s[0].isdigit():
        return False
    low = s.lower()
    if any(w in low for w in NOTE_WORDS):
        return False
    return sum(c.isalpha() for c in s) >= 2


def cname(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def ckey(s):
    return cname(s).lower()


# ---------- SUBS ----------
def corinth_subs():
    hits = list(MHP.glob("04_Marketing/**/Corinth Subcontractor*.xlsx"))
    out = []
    if not hits:
        return out
    wb = openpyxl.load_workbook(str(hits[0]), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(values_only=True):
        cells = [cname(c) for c in row if c is not None]
        if not cells or not cells[0]:
            continue
        name = cells[0]
        trade = cells[1] if len(cells) > 1 else ""
        phone = next((c for c in cells if re.search(r"\d{3}[-.\s]?\d{3,4}", c)), "")
        if phone in (trade,):
            trade = ""
        out.append({"name": name, "trade": trade[:60], "phone": phone, "src": "Corinth list"})
    return out


def sheet_subs():
    """Harvest the 'Subcontractor Name' column from every estimate sheet."""
    found = {}   # ckey -> {name, jobs:set}
    for base in ("01_Active Projects", "02_Dead Projects"):
        for xlsx in (MHP / base).rglob("*.xlsx"):
            if xlsx.name.startswith("~$"):
                continue
            try:
                wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
            except Exception:
                continue
            proj = xlsx.relative_to(MHP / base).parts[0]
            for ws in wb.worksheets:
                # the canned 'Filled In Example' sheet rides along in ~76 files — harvest the
                # names from it but DON'T count it as a real job (that's what inflated the count).
                is_example = ws.title.strip().lower() in {"filled in example", "sheet3", "example"}
                rows = list(ws.iter_rows(min_row=1, max_row=210, values_only=True))
                col = None
                for r in rows[:12]:
                    for ci, c in enumerate(r):
                        if c and "subcontractor name" in str(c).strip().lower():
                            col = ci
                    if col is not None:
                        break
                if col is None:
                    continue
                for r in rows:
                    if col < len(r) and r[col]:
                        nm = cname(r[col])
                        if nm.lower() in NON_NAMES or len(nm) < 2 or nm.replace(".", "").isdigit():
                            continue
                        rec = found.setdefault(ckey(nm), {"name": nm, "jobs": set()})
                        if not is_example:
                            rec["jobs"].add(proj)
    return found


def existing_csv_subs():
    f = MHP / "Sub_Contractor_List.csv"
    out = []
    if f.exists():
        for r in csv.DictReader(open(f)):
            out.append({"name": r["name"], "trade": r["type"],
                        "jobs": int(r.get("mention_count") or 0), "projects": r.get("projects", "")})
    return out


def build_subs(con):
    con.executescript("DROP TABLE IF EXISTS subs;"
                      "CREATE TABLE subs(name TEXT, trade TEXT, phone TEXT, jobs INTEGER,"
                      " projects TEXT, source TEXT);")
    merged = {}  # ckey -> record

    def add(name, trade="", phone="", jobs=0, projects="", source=""):
        k = ckey(name)
        if k in NON_NAMES or len(k) < 2 or not looks_like_name(name):
            return
        r = merged.setdefault(k, {"name": cname(name), "trade": "", "phone": "",
                                  "jobs": 0, "projects": "", "sources": set()})
        if trade and not r["trade"]:
            r["trade"] = trade
        if phone and not r["phone"]:
            r["phone"] = phone
        r["jobs"] = max(r["jobs"], jobs)
        if projects and not r["projects"]:
            r["projects"] = projects
        if source:
            r["sources"].add(source)

    for s in corinth_subs():
        add(s["name"], s["trade"], s["phone"], source="Corinth list")
    for k, v in sheet_subs().items():
        add(v["name"], jobs=len(v["jobs"]), projects="; ".join(sorted(v["jobs"]))[:120],
            source="estimate sheets")
    for s in existing_csv_subs():
        add(s["name"], s["trade"], jobs=s["jobs"], projects=s["projects"], source="job history")

    for r in merged.values():
        con.execute("INSERT INTO subs VALUES (?,?,?,?,?,?)",
                    (r["name"], r["trade"], r["phone"], r["jobs"], r["projects"],
                     ", ".join(sorted(r["sources"]))))
    con.commit()
    return len(merged)


# ---------- PROJECTS ----------
MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "jul": 7, "aug": 8,
          "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def parse_est_date(fn):
    """Best-effort estimate date from a filename -> 'YYYY-MM-DD' (for picking the current rev)."""
    f = fn.lower()
    m = re.search(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s*(\d{1,2})?[,\s]*\s*(20\d{2})", f)
    if m:
        return f"{int(m.group(3)):04d}-{MONTHS[m.group(1)[:3]]:02d}-{int(m.group(2) or 1):02d}"
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", f)
    if m:
        mo, day, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr += 2000 if yr < 100 else 0
        if 1 <= mo <= 12 and 1 <= day <= 31 and 2018 <= yr <= 2030:
            return f"{yr:04d}-{mo:02d}-{day:02d}"
    m = re.search(r"(\d{2})(\d{2})(20\d{2})", f)          # mmddyyyy
    if m and 1 <= int(m.group(1)) <= 12 and 1 <= int(m.group(2)) <= 31:
        return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    m = re.search(r"(20\d{2})", f)
    return f"{m.group(1)}-00-00" if m else ""


def date_estimates(con):
    try:
        con.execute("ALTER TABLE estimates ADD COLUMN est_date TEXT")
    except sqlite3.OperationalError:
        pass
    for eid, sf in con.execute("SELECT id,source_file FROM estimates").fetchall():
        con.execute("UPDATE estimates SET est_date=? WHERE id=?",
                    (parse_est_date(Path(sf).name), eid))
    con.commit()


def reclassify(con):
    date_estimates(con)
    idx = {r["project"]: r for r in csv.DictReader(open(MHP / "Project_Index.csv"))}
    for col in ("last_activity",):
        try:
            con.execute(f"ALTER TABLE projects ADD COLUMN {col} TEXT")
        except sqlite3.OperationalError:
            pass

    def bucket(latest):
        if not latest:
            return "Unknown"
        yr = latest[:4]
        if yr >= TODAY_YEAR:
            return "Active"
        if yr == "2025":
            return "Aging"
        return "Likely Done"

    for (name, pid, status) in con.execute("SELECT name,id,status FROM projects").fetchall():
        latest = (idx.get(name, {}) or {}).get("latest_date", "")
        if status == "Active":
            con.execute("UPDATE projects SET status=?, last_activity=? WHERE id=?",
                        (bucket(latest), latest, pid))
        else:
            con.execute("UPDATE projects SET last_activity=? WHERE id=?", (latest, pid))

    # add bid / paused jobs that were never in the index
    from select_corpus import slug
    for sub, st in (("_BID PIPELINE", "Bid"), ("_PAUSED", "Paused")):
        d = MHP / "01_Active Projects" / sub
        if not d.exists():
            continue
        for p in sorted(d.iterdir()):
            if p.is_dir():
                con.execute("INSERT OR REPLACE INTO projects(id,name,type,market,status,last_activity)"
                            " VALUES (?,?,?,?,?,?)",
                            (slug(p.name), p.name, "", "", st, ""))
    con.commit()


# ---------- CREW ----------
# Curated from NMHP Personnel Billing Rate Schedule (Apr 2024) + Contact Our Team.
CREW = [
    ("Rick Burge", "Operating Manager / CEO", "$165/hr", "(662) 292-1320", "admin@northmshomepros.com"),
    ("Josh Harris", "Sr. Superintendent · Licensed Contractor R21909", "$145/hr", "(662) 871-8071", "josh@northmshomepros.com"),
    ("Jason Yant", "Senior Project Manager", "$115/hr", "(662) 397-5942", "jason@northmshomepros.com"),
    ("Walt Burge", "Senior Project Coordinator", "—", "(662) 292-5533", "waltb@northmshomepros.com"),
    ("Patton Brock Burge", "Accounting / Bookkeeping", "$75/hr", "", "brockb@northmshomepros.com"),
    ("Michael Todd Murphy", "Administrative Project Support", "$65/hr", "", ""),
    ("Sandi Woods", "Materials Specialist", "—", "(951) 265-8892", "sandiw@northmshomepros.com"),
]


def build_crew(con):
    con.executescript("DROP TABLE IF EXISTS crew;"
                      "CREATE TABLE crew(name TEXT, role TEXT, rate TEXT, phone TEXT, email TEXT);")
    con.executemany("INSERT INTO crew VALUES (?,?,?,?,?)", CREW)
    con.commit()
    return len(CREW)


def main():
    con = sqlite3.connect(DB)
    n = build_subs(con)
    nc = build_crew(con)
    reclassify(con)
    from collections import Counter
    dist = dict(con.execute("SELECT status,COUNT(*) FROM projects GROUP BY status").fetchall())
    print(f"Subs: {n} real subcontractors/suppliers (was 10)")
    print(f"Crew: {nc} team members")
    print(f"Projects by status: {dist}")
    con.close()


if __name__ == "__main__":
    main()
