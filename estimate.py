"""The estimation engine — generate a new estimate from the CSI-keyed cost catalog.

Modes:
  generate  — price a scope at catalog medians -> estimate_demo.md + estimate_out.xlsx (native MHP template)
  reprice   — re-price an existing estimate's lines at today's catalog medians, show the delta

Catalog + template rows are both keyed on (CSI item #, canonical description), so lookups are robust
to whitespace/case and disambiguate material/labor pairs that share a CSI code. Unit-rate items price
qty*median; lump-sum items (General Conditions, cabinets) price from the lump catalog.
"""
import sqlite3
import sys
from pathlib import Path

import openpyxl

from normalize import canon, norm_unit

DB = Path(__file__).parent / "mhp.db"
HERE = Path(__file__).parent
BLANK = HERE / "mhp_template_blank.xlsx"
DEFAULT_MARKUP = 1.18   # 17.8% median observed across 4,251 lines


def load_catalog(con):
    """A description can have several unit variants (sqft vs a few stray 'each'). Keep the
    DOMINANT one — most jobs — so a 3-job outlier rate can't hijack a 68-job line."""
    unit, lump = {}, {}
    for ino, cd, div, unit_, j, med, p25, p75 in con.execute(
        "SELECT item_no,canon_desc,division,unit,n_jobs,median_unit_price,p25,p75 FROM unit_costs "
        "ORDER BY n_jobs DESC"):
        if cd in unit:
            continue   # already have the higher-job variant
        unit[cd] = {"item_no": ino, "division": div, "unit": unit_, "jobs": j,
                    "median": med, "p25": p25, "p75": p75}
    for ino, cd, div, j, med, p25, p75 in con.execute(
        "SELECT item_no,canon_desc,division,n_jobs,median_total,p25,p75 FROM lump_costs "
        "ORDER BY n_jobs DESC"):
        if cd in lump:
            continue
        lump[cd] = {"item_no": ino, "division": div, "jobs": j, "median": med, "p25": p25, "p75": p75}
    return unit, lump


def contingency(p25, p75, med):
    if not med:
        return 0
    spread = (p75 - p25) / med
    return 5 if spread < 0.15 else 10 if spread < 0.4 else 20


def price_scope(unit, lump, scope, markup=DEFAULT_MARKUP):
    lines, missing = [], []
    for s in scope:
        cd = canon(s["description"])
        qty = s.get("qty", 1)
        u, lp = unit.get(cd), lump.get(cd)
        # When an item exists as both a unit rate and a whole-job lump, trust the
        # better-backed one (more jobs). Without this, a $175 per-each "plumbing
        # fixtures" allowance beats the 62-job $2,350 whole-house lump.
        if u is not None and (lp is None or u["jobs"] >= lp["jobs"]):
            h = u
            item_total = round(h["median"] * qty, 2)
            lines.append({"kind": "unit", "item_no": h["item_no"], "division": h["division"],
                          "description": s["description"], "unit": h["unit"], "qty": qty,
                          "rate": h["median"], "item_total": item_total,
                          "sov_total": round(item_total * markup, 2), "jobs": h["jobs"],
                          "p25": h["p25"], "p75": h["p75"],
                          "contingency": contingency(h["p25"], h["p75"], h["median"])})
        elif lp is not None:
            h = lp
            item_total = round(h["median"] * qty, 2)
            lines.append({"kind": "lump", "item_no": h["item_no"], "division": h["division"],
                          "description": s["description"], "unit": "lump", "qty": qty,
                          "rate": h["median"], "item_total": item_total,
                          "sov_total": round(item_total * markup, 2), "jobs": h["jobs"],
                          "p25": h["p25"], "p75": h["p75"],
                          "contingency": contingency(h["p25"], h["p75"], h["median"])})
        else:
            missing.append(s)
    return lines, missing


def write_xlsx(lines, out_path):
    """Fill a copy of the blank MHP template; Excel recomputes Item Total/Markup/SOV on open."""
    wb = openpyxl.load_workbook(str(BLANK), data_only=False)
    ws = wb["Template"]
    rowmap = {}   # (canon_desc) -> row  (canon desc is unique per template row)
    for r in range(7, 183):
        b = ws.cell(row=r, column=2).value
        if b and not str(b).strip().lower().startswith("division"):
            rowmap[canon(b)] = r
    mapped = unmapped = 0
    for ln in lines:
        r = rowmap.get(canon(ln["description"]))
        if r is None:
            unmapped += 1
            continue
        ws.cell(row=r, column=3).value = ln["qty"]      # C = qty
        ws.cell(row=r, column=5).value = ln["rate"]     # E = unit price (or lump total at qty 1)
        mapped += 1
    wb.save(str(out_path))
    return mapped, unmapped


def render(title, lines, missing, markup):
    sub = sum(l["item_total"] for l in lines)
    sov = sum(l["sov_total"] for l in lines)
    cont = sum(l["item_total"] * l["contingency"] / 100 for l in lines)
    out = [f"# Estimate — {title}", "",
           f"Priced from CSI-keyed catalog medians · markup {(markup-1)*100:.0f}% · "
           f"{len(lines)} lines ({sum(1 for l in lines if l['kind']=='lump')} lump), "
           f"{len(missing)} unpriced", "",
           "| CSI | Division | Item | Qty | Unit | Rate | Item Total | Backed | Band | Cont |",
           "|---|---|---|--:|--|--:|--:|--:|--|--:|"]
    for l in sorted(lines, key=lambda x: x["division"] or ""):
        d = (l["division"] or "").replace("Division ", "Div ")[:16]
        out.append(f"| {l['item_no'] or ''} | {d} | {l['description'][:24]} | {l['qty']:g} | "
                   f"{l['unit'] or ''} | ${l['rate']:,.2f} | ${l['item_total']:,.0f} | {l['jobs']}j | "
                   f"${l['p25']:,.2f}-${l['p75']:,.2f} | {l['contingency']}% |")
    out += ["", f"**Subtotal (cost):** ${sub:,.0f}",
            f"**Suggested contingency:** ${cont:,.0f}",
            f"**Bid (with {(markup-1)*100:.0f}% markup):** ${sov:,.0f}", ""]
    if missing:
        out += ["**Unpriced (no history — manual rate / invoice data needed):**"]
        out += [f"- {m['description']} ({m.get('qty',1):g})" for m in missing]
    return "\n".join(out) + "\n"


KITCHEN_SCOPE = [
    {"description": "General Conditions", "qty": 1},
    {"description": "Project Coordination (Supervision)", "qty": 2},
    {"description": "Structure Demolition", "qty": 1},
    {"description": "Framing Material", "qty": 300},
    {"description": "Framing Labor", "qty": 300},
    {"description": "Drywall ", "qty": 850},
    {"description": "Interior Trim Material", "qty": 300},
    {"description": "Interior Trim Labor", "qty": 300},
    {"description": "LVT Flooring - Materials", "qty": 300},
    {"description": "LVT Flooring - Labor", "qty": 300},
    {"description": "Interior Paint Material", "qty": 850},
    {"description": "Interior Paint Labor", "qty": 850},
    {"description": "Electrical Material", "qty": 300},
    {"description": "Electrical Labor", "qty": 300},
    {"description": "Plumbing Material & Labor", "qty": 3},
    {"description": "Kitchen Cabinets", "qty": 70},
    {"description": "Countertop Material", "qty": 60},
    {"description": "Countertop Labor", "qty": 60},
    {"description": "Backsplash Tile Materails", "qty": 40},
    {"description": "Backsplash Tile Labor", "qty": 40},
]


def reprice(con, unit, like):
    row = con.execute("SELECT id,source_file FROM estimates WHERE source_file LIKE ? "
                      "AND parse_confidence!='FAILED' LIMIT 1", (f"%{like}%",)).fetchone()
    if not row:
        return f"No estimate matching '{like}'."
    eid, sf = row
    items = con.execute("""SELECT description,canon_desc,qty,unit_price,item_total FROM line_items
                           WHERE estimate_id=? AND price_kind='UNIT_RATE' AND qty>0""", (eid,)).fetchall()
    out = [f"# Reprice check — {sf.split('/')[-1]}", "",
           "Original bid rate vs catalog median. Δ shows where the job priced above/below MHP norm.", "",
           "| Item | Qty | Bid rate | Catalog med | Δ | Bid total | At catalog |",
           "|---|--:|--:|--:|--:|--:|--:|"]
    bid_sum = cat_sum = 0
    for desc, cd, qty, up, it in items:
        h = unit.get(cd)
        if not h:
            continue
        med = h["median"]
        cat_total = med * qty
        bid_sum += it or 0
        cat_sum += cat_total
        d = (up - med) / med * 100 if med else 0
        out.append(f"| {desc[:26]} | {qty:g} | ${up:,.2f} | ${med:,.2f} | {d:+.0f}% | "
                   f"${it or 0:,.0f} | ${cat_total:,.0f} |")
    delta = (bid_sum - cat_sum) / cat_sum * 100 if cat_sum else 0
    out += ["", f"**Job unit-rate cost:** ${bid_sum:,.0f} · **At catalog medians:** ${cat_sum:,.0f} · "
            f"**priced {delta:+.0f}% vs norm**"]
    return "\n".join(out) + "\n"


def main():
    con = sqlite3.connect(DB)
    unit, lump = load_catalog(con)
    mode = sys.argv[1] if len(sys.argv) > 1 else "generate"
    if mode == "reprice":
        like = sys.argv[2] if len(sys.argv) > 2 else "Mason Kitchen Estimate 11-10-25"
        text = reprice(con, unit, like)
        (HERE / "reprice_demo.md").write_text(text)
        print(text)
    else:
        lines, missing = price_scope(unit, lump, KITCHEN_SCOPE)
        text = render("Kitchen Remodel (~300 sqft, demo scope)", lines, missing, DEFAULT_MARKUP)
        (HERE / "estimate_demo.md").write_text(text)
        mapped, unmapped = write_xlsx(lines, HERE / "estimate_out.xlsx")
        print(text)
        print(f"\nxlsx: wrote estimate_out.xlsx — {mapped} lines mapped to template rows, {unmapped} unmapped")
    con.close()


if __name__ == "__main__":
    main()
