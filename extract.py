"""Layer 0 extractor — parse MHP estimate spreadsheets into structured line items.

Replaces the old `biggest_dollar_in_xlsx` guess with a real templated parse:
MHP estimates on a standardized CSI-division template (header row with
Item #/Description/Unit/.../Item Total/SOV Total). We detect that structure,
extract every priced line, and FLAG anything we don't fully trust.
"""
import re
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

from select_corpus import build_corpus, slug

DB = Path(__file__).parent / "mhp.db"
SCHEMA = Path(__file__).parent / "schema.sql"
EXAMPLE_SHEETS = {"filled in example", "sheet3", "example"}

ERR = {"#div/0!", "#ref!", "#value!", "#n/a", "#name?", "#null!", "#num!"}


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip().lower() if v is not None else ""


# Aggregate roll-up rows that ride along in the template but are NOT line items.
AGGREGATE_ROWS = {"all subtotals", "subtotal", "subtotals", "grand total"}


def num(v):
    """Coerce a cell to float, or None for blanks/formula-errors/text."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ERR:
        return None
    s = s.replace("$", "").replace(",", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def market_of(folder_name: str) -> str:
    n = folder_name.lower()
    pick = ("eaton", "buchanan", "knight", "cromeans", "pickwick", "corinth")
    if any(k in n for k in pick):
        return "Pickwick"
    return "Oxford"


# Column header -> canonical field. Matched against normalized header text.
HEADER_MAP = [
    ("item_no", ["item #", "item#", "item number"]),
    ("description", ["description"]),
    ("qty", ["approx. quantity", "approx quantity", "quantity"]),
    ("unit", ["unit of measure"]),
    ("unit_price", ["unit price"]),
    ("material", ["material & equipment", "material and equipment", "material"]),
    ("labor", ["self performed labor", "labor"]),
    ("sub_bid", ["subcontractor bid"]),
    ("item_total", ["item total"]),
    ("markup", ["markup"]),
    ("sov_total", ["sov total"]),
    ("sub_name", ["subcontractor name"]),
]


def find_header(ws):
    """Return (header_row_idx, {field: col_idx}) or (None, None)."""
    rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
    for i, row in enumerate(rows):
        cells = [norm(c) for c in row]
        joined = " ".join(cells)
        if "item" in joined and "description" in joined and (
            "unit of measure" in joined or "quantity" in joined
        ):
            colmap = {}
            for ci, txt in enumerate(cells):
                for field, aliases in HEADER_MAP:
                    if field in colmap:
                        continue
                    if any(a == txt or a in txt for a in aliases):
                        colmap[field] = ci
            if "description" in colmap and "item_total" in colmap:
                return i, colmap
    return None, None


def pick_sheet(wb):
    """Prefer 'Template'; never fall through to the canned example sheet."""
    names = wb.sheetnames
    ordered = [n for n in names if norm(n) == "template"]
    ordered += [n for n in names if norm(n) not in EXAMPLE_SHEETS and n not in ordered]
    return ordered


def parse_estimate(path: Path):
    """Return dict with line_items + totals + flags, or {'flags':['FAILED:...']}."""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    for sheet_name in pick_sheet(wb):
        ws = wb[sheet_name]
        hdr, cols = find_header(ws)
        if hdr is None:
            continue
        items = []
        division = None
        for row in ws.iter_rows(min_row=hdr + 2, max_row=hdr + 400, values_only=True):
            def cell(field):
                ci = cols.get(field)
                return row[ci] if ci is not None and ci < len(row) else None

            desc_raw = cell("description")
            desc = norm(desc_raw)
            if desc.startswith("division"):
                division = str(desc_raw).strip()
                continue
            if desc in AGGREGATE_ROWS:
                continue  # sheet's own roll-up ('All Subtotals') — equals Σ of real lines,
                          # so ingesting it as a line item double-counted the whole bid.
            it = num(cell("item_total"))
            if it is None or it <= 0 or not desc:
                continue  # unused catalog row or division subtotal
            items.append({
                "division": division,
                "item_no": (str(cell("item_no")).strip() if cell("item_no") else None),
                "description": str(desc_raw).strip(),
                "qty": num(cell("qty")),
                "unit": (str(cell("unit")).strip() if cell("unit") else None),
                "unit_price": num(cell("unit_price")),
                "material": num(cell("material")),
                "labor": num(cell("labor")),
                "sub_bid": num(cell("sub_bid")),
                "item_total": it,
                "sov_total": num(cell("sov_total")),
                "sub_name": (str(cell("sub_name")).strip() if cell("sub_name") else None),
            })
        items = _dedup(items)
        if items:
            return {"sheet": sheet_name, "items": items}
        return {"sheet": sheet_name, "items": [], "no_items": True}
    return {"failed": "NO_TEMPLATE_HEADER"}


def _dedup(items):
    """Drop exact-duplicate line items within a sheet.

    Some source sheets carry a duplicated section (e.g. Eaton bb-3086 had 30 repeated
    lines, ~$107K of phantom total). A line keyed identical on item#/desc/qty/price/total
    is a copy-paste artifact, not two real scopes — keep the first, drop the rest.
    """
    seen, out = set(), []
    for it in items:
        key = (it["item_no"], (it["description"] or "").strip().lower(),
               it["qty"], it["unit_price"], it["item_total"])
        if key in seen:
            continue
        seen.add(key)
        out.append(it)
    return out


def main():
    con = sqlite3.connect(DB)
    con.executescript("DROP TABLE IF EXISTS projects; DROP TABLE IF EXISTS estimates;"
                      "DROP TABLE IF EXISTS line_items; DROP TABLE IF EXISTS actuals;")
    con.executescript(SCHEMA.read_text())

    kitchen_only = "--kitchen" in sys.argv
    corpus = build_corpus(kitchen_only=kitchen_only)
    scope = "kitchen" if kitchen_only else "FULL portfolio"
    print(f"Extracting {scope}: {len(corpus)} projects...")
    rev_groups = defaultdict(list)   # old-slug base -> [(eid, rel_lower)] : revision siblings of one estimate
    for entry in corpus:
        pr = entry["project"]
        pid = slug(pr["project"])
        con.execute("INSERT OR REPLACE INTO projects VALUES (?,?,?,?,?)",
                    (pid, pr["project"], pr["type"], market_of(pr["project"]), pr["status"]))

        for est in entry["estimates"]:
            rel = str(est.relative_to(entry["dir"]))
            eid = slug(rel)
            flags = []
            if re.search(r"phase", est.name, re.I):
                flags.append("PHASE_ONLY")
            if "_onedrive4" in est.name.lower():
                flags.append("DUPLICATE_EXPORT")

            try:
                res = parse_estimate(est)
            except Exception as e:
                res = {"failed": f"EXC:{type(e).__name__}"}

            if res.get("failed"):
                conf, flags = "FAILED", flags + [res["failed"]]
                con.execute("INSERT OR REPLACE INTO estimates VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                            (eid, pid, str(est), "xlsx", None, 0, 0, 0, None, conf, ",".join(flags)))
                continue
            if res.get("no_items"):
                conf, flags = "FAILED", flags + ["NO_LINE_ITEMS"]
                con.execute("INSERT OR REPLACE INTO estimates VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                            (eid, pid, str(est), "xlsx", res["sheet"], 0, 0, 0, None, conf, ",".join(flags)))
                continue

            items = res["items"]
            sum_item = round(sum(i["item_total"] or 0 for i in items), 2)
            sum_sov = round(sum(i["sov_total"] or 0 for i in items), 2)
            conf = "FLAGGED" if flags else "CLEAN"
            con.execute("INSERT OR REPLACE INTO estimates VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (eid, pid, str(est), "xlsx", res["sheet"], len(items),
                         sum_item, sum_sov, None, conf, ",".join(flags)))
            rev_groups[re.sub(r"[^a-z0-9]+", "-", rel.lower()).strip("-")[:80]].append((eid, rel.lower()))
            for it in items:
                con.execute(
                    "INSERT INTO line_items (estimate_id,division,item_no,description,qty,unit,"
                    "unit_price,material,labor,sub_bid,item_total,sov_total,sub_name) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (eid, it["division"], it["item_no"], it["description"], it["qty"], it["unit"],
                     it["unit_price"], it["material"], it["labor"], it["sub_bid"],
                     it["item_total"], it["sov_total"], it["sub_name"]))

        # actuals (materials closeouts — partial; biggest labeled total best-effort)
        for cl in entry["closings"]:
            try:
                total = closing_total(cl)
            except Exception:
                total = None
            con.execute("INSERT INTO actuals VALUES (?,?,?)", (pid, str(cl), total))

    # Revision siblings: same estimate re-saved (RV1/RV2/working-doc/later date). They share the
    # pre-truncation slug. Keep the latest (lexicographically-greatest name — the RV/date suffix
    # sorts last) and flag the rest SUPERSEDED so one job isn't double-weighted in the catalog.
    def _rev_rank(rel):
        # A revision APPENDS tokens (RV1, 'working doc', a later date), so the newer file's
        # name is longer. Rank by length (ext stripped), lexicographic tiebreak (rv2 > rv1).
        base = re.sub(r"\.[a-z0-9]+$", "", rel)
        return (len(base), base)

    n_superseded = 0
    for members in rev_groups.values():
        if len(members) < 2:
            continue
        keeper = max(members, key=lambda m: _rev_rank(m[1]))[0]
        for eid, _ in members:
            if eid != keeper:
                con.execute(
                    "UPDATE estimates SET flags=TRIM(COALESCE(flags,'')||',SUPERSEDED', ','), "
                    "parse_confidence='SUPERSEDED' WHERE id=?", (eid,))
                n_superseded += 1

    con.commit()
    n_est = con.execute("SELECT COUNT(*) FROM estimates").fetchone()[0]
    n_li = con.execute("SELECT COUNT(*) FROM line_items").fetchone()[0]
    print(f"Wrote {n_est} estimates, {n_li} line items "
          f"({n_superseded} superseded revisions) -> {DB.name}")
    con.close()


def closing_total(path: Path):
    """Best-effort total from a closing/materials sheet (labeled 'total' row)."""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    best = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [norm(c) for c in row]
            if any("total" in c for c in cells):
                for c in row:
                    v = num(c)
                    if v and (best is None or v > best):
                        best = v
    return best


if __name__ == "__main__":
    main()
